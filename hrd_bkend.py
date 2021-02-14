import requests
import re
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Side, Border
import os

from bs4 import BeautifulSoup as bs

RE_COMP = re.compile('[b\\/:*?"<>|\\t]')
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DL_DIR = os.path.join(BASE_DIR, "result")
EXCEL_BASE_URL = "http://www.hrd.go.kr/hrdp/ps/ppsmo/excelDownAll0109P.do?"
FILE_DIR = os.path.join(DL_DIR, f"result.xlsx")

HEADERS = {
    "Accept": "text/html,application/xhtml+xml,"
              "application/xml;q=0.9,image/avif,"
              "image/webp,image/apng,*/*;q=0.8,"
              "application/signed-exchange;v=b3;q=0.9",
    "Accept-Encoding": "gzip, deflate",
    "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7,ja;q=0.6",
    "Cache-Control": "max-age=0",
    "Connection": "keep-alive",
    "DNT": "1",
    "Host": "www.hrd.go.kr",
    "Upgrade-Insecure-Requests": "1",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/86.0.4240.198 Safari/537.36"
}

if not os.path.isdir(DL_DIR):
    os.mkdir(DL_DIR)


def getDetail(session, keyword, cCode1, cCode2, ncsCode, areaCode, startDate, endDate):
    url = 'https://app.hrd.go.kr/hrdp/ti/ptiao/PTIAO0100L.do?'
    payload = ''
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "result"

    if ncsCode is not None:
        payload += f"ncs={ncsCode}"
    payload += f'crseTraceseSe{cCode1}={cCode2}&area={areaCode}&pageId={cCode1}&' \
               'bgrlInstYn=&currentTab=2&' \
               'pageIndex=1&pageSize=10&' \
               'orderKey=2&orderBy=ASC&' \
               'mberId=&mberSe=&pop=&'

    if keyword != '':
        payload += f'keyword2=&keyword1=&keyword={keyword}&keywordType=1&'
    else:
        payload += f'keyword2=&keyword1=&keyword=&keywordType=1&'

    payload += f"startDate={startDate}&endDate={endDate}&" \
               f"i2=A&dghtSe=A&traingMthCd=A&totamtSuptYn=A&" \
               f"totTraingTime=A&intrstInfoAdiCn=&pageOrder=2ASC&pageRow=100&"

    req = session.get(url+payload)
    reqbs = bs(req.content, "html.parser")
    pages = int(re.sub('[(),]', '', reqbs.select('span.count')[1].text)) // 100 + 2
    print(f"count : {reqbs.select('span.count')[1].text}, pages : {pages-1}")

    for i in range(1, pages):
        print("page : ", i)

        urlWithPayload = re.sub('pageIndex=\\d{1,2}', f'pageIndex={i}', url + payload)

        getContent = requests.get(urlWithPayload).content
        getContentbs = bs(getContent, "html.parser")
        detail_List = getContentbs.select('#searchForm > div > div.contentArea.section-training > div.detailListWrap > ul '
                                          '> li')
        for j in detail_List:
            tmp = []

            title = re.sub('\\s', '', j.select('div.title > a > p')[0].text)
            subtitle = re.sub('\\s', '', j.select('div.content > p')[0].text)
            res = re.findall('[A-Z]{3}\\d{14}|\\d{12}|C\\d{4}|\\d{1,2}', j.select('a')[0]['onclick'])

            if len(res) != 4:
                detail = 'https://m.hrd.go.kr/hrdp/co/pcobo/PCOBO0100P.do?' \
                         f'tracseId={res[0]}&' \
                         f'tracseTme={res[1]}&' \
                         f'crseTracseSe={res[4]}&' \
                         f'trainstCstmrId={res[3]}'

            else:
                detail = 'https://m.hrd.go.kr/hrdp/co/pcobo/PCOBO0100P.do?' \
                         f'tracseId={res[0]}&' \
                         f'tracseTme={res[1]}&' \
                         f'crseTracseSe={res[3]}&' \
                         f'trainstCstmrId={res[2]}'

            detail_req = requests.get(detail)
            detail_req_bs = bs(detail_req.content, "html.parser")

            print(title, subtitle, detail)

            # 업체명
            tmp.append(re.sub('\\s', '', detail_req_bs.find('p', {'class': 'add'}).text))

            # 과정명
            tmp.append(re.sub('\\s|모집마감', '', detail_req_bs.find('h4', {'class': 'tit'}).text))

            # n회차 텍스트를 제외한 훈련기간
            tmp.append(re.sub('\\s|\(\\d회차\)', '', detail_req_bs.select('li:nth-child(6) > span.con')[0].text))

            # 최근회차
            training_time = detail_req_bs.select_one('li:nth-child(6) > span.con').text
            recent_year = int(re.findall('\\d{4}', training_time)[0])
            recent_Time = int(re.findall('(\\d)', training_time)[0])
            recent = int(re.sub('회차|~|\\s|[\\(\\)]|\\d{4}-\\d{2}-\\d{2}', '', training_time)[0]) - 1

            gap_of_list = recent - len(detail_req_bs.select(f"dl tbody > tr:nth-child(2) > td")) - 2

            if recent == 0:
                recent = 1

            # 훈련시간
            tmp.append(re.sub('\\d*일|,|총|시간', '', detail_req_bs.select('li:nth-child(7) > span.con')[0].text))

            # 수강인원
            try:
                int(detail_req_bs.select(f"dl:nth-child({recent}) tbody > tr:nth-child(2) > td")[0].text)

                tmp.append(
                    detail_req_bs.select(f"dl:nth-child({recent}) tbody > tr:nth-child(2) > td")[0].text
                )

                # 정원(모집인원)
                tmp.append(detail_req_bs.select(f"dl:nth-child({recent}) tbody > tr > td")[0].text)

            except ValueError:
                sep_num = re.findall(
                    '\\d{1,3}',
                    detail_req_bs.select(f"dl:nth-child({recent}) tbody > tr:nth-child(2) > td")[0].text
                )

                tmp.append(sep_num[0]) if sep_num[0] > sep_num[1] else tmp.append(sep_num[1])

                # 정원(모집인원)
                tmp.append(detail_req_bs.select(f"dl:nth-child({recent}) tbody > tr > td")[0].text)

            except IndexError:
                try:
                    int(detail_req_bs.select(f"dl:nth-child({gap_of_list}) tbody > tr:nth-child(2) > td")[0].text)

                    tmp.append(
                        detail_req_bs.select(f"dl:nth-child({gap_of_list}) tbody > tr:nth-child(2) > td")[0].text
                    )
                    tmp.append(detail_req_bs.select(f"dl:nth-child({gap_of_list}) tbody > tr > td")[0].text)

                except ValueError:
                    sep_num = re.findall(
                        '\\d{1,3}',
                        detail_req_bs.select(f"dl:nth-child({gap_of_list}) tbody > tr:nth-child(2) > td")[0].text
                    )

                    tmp.append(sep_num[0]) if sep_num[0] > sep_num[1] else tmp.append(sep_num[1])

                    # 정원(모집인원)
                    tmp.append(detail_req_bs.select(f"dl:nth-child({gap_of_list}) tbody > tr > td")[0].text)

            # 취업률
            try:
                if len(detail_req_bs.select(f"dl:nth-child({recent}) tbody > tr > td")) != 4:
                    employee_percent = detail_req_bs.select(f"dl:nth-child({recent}) tbody > tr > td")
                    tmp.append(re.sub('\\s|\\(\\d{1,3}/\\d{1,3}\\)', '', employee_percent[3].text))
                    tmp.append(re.sub('\\s|\\(\\d{1,3}/\\d{1,3}\\)', '', employee_percent[4].text))
                    tmp.append(re.sub('\\s|\\(\\d{1,3}/\\d{1,3}\\)', '', employee_percent[5].text))

            except IndexError:
                if len(detail_req_bs.select(f"dl:nth-child({gap_of_list}) tbody > tr > td")) != 4:
                    employee_percent = detail_req_bs.select(f"dl:nth-child({gap_of_list}) tbody > tr > td")
                    tmp.append(re.sub('\\s|\\(\\d{1,3}/\\d{1,3}\\)', '', employee_percent[3].text))
                    tmp.append(re.sub('\\s|\\(\\d{1,3}/\\d{1,3}\\)', '', employee_percent[4].text))
                    tmp.append(re.sub('\\s|\\(\\d{1,3}/\\d{1,3}\\)', '', employee_percent[5].text))

            ws.append(tmp)
            wb.save(filename=os.path.join(DL_DIR, "result.xlsx"))
            getExcel(session, title, subtitle, res[0], recent_year, recent_Time)

    os.startfile(DL_DIR)


def getExcel(sess, title, subtitle, tracseId, recentYear, recentTme):
    args = f"tracseId={tracseId}&ncsYn=Y&pssrpYear={recentYear}&pssrpTme={recentTme}&" \
           f"mainTracseSe=&ncsAbluitFactorUnitSe=&tracseSttusCd=&jdgmnSe=&excelAllYn=&excelGbn=&A2Gbn="

    title = re.sub(RE_COMP, "", title)
    subtitle = re.sub(RE_COMP, "", subtitle)

    t = sess.get(EXCEL_BASE_URL + args)
    if t.headers.get('content-type') == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        print("Downloading... ", end=' ')
        FILE_DIR = os.path.join(DL_DIR, f"{title}.xlsx")
        with open(f'{FILE_DIR}', 'wb') as f:
            f.write(t.content)

        wb = openpyxl.load_workbook(FILE_DIR)
        ws = wb[wb.sheetnames[0]]
        set_template(ws, subtitle)
        wb.copy_worksheet(wb[wb.sheetnames[0]])
        wb[wb.sheetnames[1]].delete_cols(12)
        wb.save(os.path.join(DL_DIR, f'{title}_{subtitle}.xlsx'))
        os.remove(FILE_DIR)


def set_template(ws, acaName):
    max_rows = 0
    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )

    align = Alignment(horizontal='center', vertical='center', wrapText=True)

    ws.merge_cells('A1:L1')
    ws.merge_cells('A2:L2')
    ws['A1'] = '학원명'
    ws['A2'] = acaName
    ws['A1'].fill = PatternFill(start_color='FFFF00',
                                end_color='FFFF00',
                                fill_type='solid')

    for _ in ws.rows:
        max_rows += 1

    rows = ws.iter_rows(1, max_rows)
    for row in rows:
        for cell in row:
            cell.border = border
            cell.alignment = align

    for rows in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=12):
        for cell in rows:
            cell.fill = PatternFill(start_color='FFFF00',
                                    end_color='FFFF00',
                                    fill_type='solid')
